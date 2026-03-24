import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
)
from reportlab.lib.styles import getSampleStyleSheet

def export_to_excel(drives: list, expenses: list) -> io.BytesIO:
    wb = openpyxl.Workbook()

    # Drive Logs sheet
    ws = wb.active
    ws.title = 'Drive Logs'
    _headers(ws, ['Date','Start KM','End KM','Distance (km)','Purpose','Notes'], '2563EB')
    for d in drives:
        ws.append([
            d.get('log_date',''),
            d.get('start_km',''),
            d.get('end_km', '—'),
            d.get('distance_km', 'Incomplete'),
            d.get('purpose',''),
            d.get('notes','')
        ])
    _auto_width(ws)

    # Expenses sheet
    ws2 = wb.create_sheet('Expenses')
    _headers(ws2, ['Date','Type','Vendor','Amount','HST','Total','Receipt'], '16A34A')
    for e in expenses:
        ws2.append([
            e.get('expense_date',''),
            e.get('expense_type',''),
            e.get('vendor',''),
            float(e.get('amount', 0)),
            float(e.get('hst', 0)),
            float(e.get('total_amount', 0)),
            'Yes' if e.get('receipt_file') else ''
        ])
    _auto_width(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def _headers(ws, cols: list, color: str) -> None:
    fill = PatternFill('solid', fgColor=color)
    bold = Font(bold=True, color='FFFFFF')
    for i, h in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font, c.fill = bold, fill
        c.alignment = Alignment(horizontal='center')

def _auto_width(ws) -> None:
    for col in ws.columns:
        w = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, 40)

def export_to_pdf(drives: list, expenses: list,
                  user_name: str = '') -> io.BytesIO:
    buf   = io.BytesIO()
    doc   = SimpleDocTemplate(buf, pagesize=letter,
                leftMargin=0.75*inch, rightMargin=0.75*inch,
                topMargin=0.75*inch,  bottomMargin=0.75*inch)
    styles = getSampleStyleSheet()
    story  = []

    title = 'TrackWise Report'
    if user_name:
        title += ' - ' + user_name
    story.append(Paragraph(title, styles['Title']))
    story.append(Spacer(1, 0.2*inch))

    # Drive Logs section
    story.append(Paragraph('Drive Logs', styles['Heading2']))
    story.append(Spacer(1, 0.1*inch))
    drive_rows = [['Date', 'Start KM', 'End KM', 'Distance', 'Purpose']]
    total_km = 0
    for d in drives:
        dist = d.get('distance_km')
        if dist:
            total_km += dist
            dist_str = str(dist) + ' km'
        else:
            dist_str = 'Incomplete'
        drive_rows.append([
            d.get('log_date', ''),
            str(d.get('start_km', '')),
            str(d.get('end_km', '-')),
            dist_str,
            d.get('purpose', ''),
        ])
    drive_rows.append(['', '', 'TOTAL', str(round(total_km, 1)) + ' km', ''])
    story.append(_pdf_table(drive_rows))
    story.append(Spacer(1, 0.3*inch))

    # Expenses section
    story.append(Paragraph('Expenses', styles['Heading2']))
    story.append(Spacer(1, 0.1*inch))
    exp_rows = [['Date', 'Type', 'Vendor', 'HST', 'Total']]
    total_spent = 0
    for e in expenses:
        total_spent += float(e.get('total_amount', 0))
        exp_rows.append([
            e.get('expense_date', ''),
            e.get('expense_type', ''),
            e.get('vendor', ''),
            '$' + str(round(float(e.get('hst', 0)), 2)),
            '$' + str(round(float(e.get('total_amount', 0)), 2)),
        ])
    exp_rows.append(['', '', 'TOTAL', '', '$' + str(round(total_spent, 2))])
    story.append(_pdf_table(exp_rows))

    doc.build(story)
    buf.seek(0)
    return buf

def _pdf_table(data: list) -> Table:
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0),  (-1, 0),  colors.HexColor('#2563EB')),
        ('TEXTCOLOR',  (0, 0),  (-1, 0),  colors.white),
        ('FONTNAME',   (0, 0),  (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0),  (-1, -1), 9),
        ('GRID',       (0, 0),  (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2),
            [colors.white, colors.HexColor('#f8fafc')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f1f5f9')),
        ('FONTNAME',   (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN',      (0, 0),  (-1, -1), 'LEFT'),
        ('PADDING',    (0, 0),  (-1, -1), 6),
    ]))
    return t